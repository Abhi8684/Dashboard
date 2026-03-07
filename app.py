import os
import uuid
from flask import Flask, render_template, request, redirect, url_for, session
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

# ── Dash app (must be imported after Flask so it can mount inside Flask) ──
from dashboard import create_dash_app

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
ALLOWED_EXTENSIONS = {"xlsx", "xls"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Flask server — Dash will mount itself onto this
flask_app = Flask(__name__)
flask_app.secret_key = os.environ.get("SECRET_KEY", "kannetn-dashboard-secret-2026")
flask_app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
flask_app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16 MB

# Mount Dash dashboard at /dashboard/
dash_app = create_dash_app(flask_app)

# ── Expose the WSGI server object (for gunicorn) ──
server = flask_app


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@flask_app.route("/", methods=["GET"])
def index():
    return render_template("upload.html")


@flask_app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return render_template("upload.html", error="No file selected.")

    file = request.files["file"]

    if file.filename == "":
        return render_template("upload.html", error="No file selected.")

    if not allowed_file(file.filename):
        return render_template(
            "upload.html", error="Invalid file type. Please upload an .xlsx or .xls file."
        )

    # Save with a unique name to avoid collisions in production
    unique_name = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
    filepath = os.path.join(flask_app.config["UPLOAD_FOLDER"], unique_name)
    file.save(filepath)

    # ── Multi-sheet validation ──────────────────────────────────────────
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception:
        os.remove(filepath)
        return render_template(
            "upload.html",
            error="Could not read the Excel file. Please upload a valid .xlsx file.",
        )

    if len(sheet_names) > 1:
        if "Daily_Activity_Tracker" not in sheet_names:
            os.remove(filepath)
            return render_template(
                "upload.html",
                error=(
                    "The uploaded file contains multiple sheets but none is named "
                    "'Daily_Activity_Tracker'. Please upload a single-sheet Excel file "
                    "or ensure it contains a sheet named 'Daily_Activity_Tracker'."
                ),
            )
        # Multiple sheets present but the required one exists — store the target sheet
        session["target_sheet"] = "Daily_Activity_Tracker"
    else:
        session.pop("target_sheet", None)

    # Store path in session so the Dash app can read it
    session["excel_path"] = filepath

    return redirect("/dashboard/")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    flask_app.run(host="0.0.0.0", port=port, debug=False)
